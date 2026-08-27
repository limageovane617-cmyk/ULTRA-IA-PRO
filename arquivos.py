# ============================================================
# 📄 ALEX IA ULTRA — SISTEMA DE ARQUIVOS
# Criada por Geovani
# ============================================================

from io import BytesIO
from pathlib import Path
import csv
import json
import zipfile

import PyPDF2
from docx import Document

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ============================================================
# 🛡️ LIMITES DE SEGURANÇA
# ============================================================

# Quantidade máxima de arquivos que o ZIP poderá processar.
MAX_ARQUIVOS_ZIP = 200

# Tamanho máximo de um arquivo interno que será lido.
MAX_ARQUIVO_INTERNO_ZIP = 5 * 1024 * 1024  # 5 MB

# Quantidade máxima total de dados internos que serão lidos.
MAX_TOTAL_ZIP = 20 * 1024 * 1024  # 20 MB


# ============================================================
# 📝 TXT
# ============================================================

def ler_txt(arquivo):
    """Lê um arquivo TXT."""

    try:
        conteudo = arquivo.read()

        if isinstance(conteudo, bytes):
            conteudo = conteudo.decode("utf-8", errors="ignore")

        return conteudo, None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 📕 PDF
# ============================================================

def ler_pdf(arquivo):
    """Extrai texto de um arquivo PDF."""

    try:
        leitor = PyPDF2.PdfReader(arquivo)

        paginas = []

        for pagina in leitor.pages:
            texto = pagina.extract_text()

            if texto:
                paginas.append(texto)

        return "\n\n".join(paginas), None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 📘 DOCX
# ============================================================

def ler_docx(arquivo):
    """Extrai texto de um arquivo DOCX."""

    try:
        documento = Document(BytesIO(arquivo.read()))

        paragrafos = []

        for paragrafo in documento.paragraphs:

            if paragrafo.text.strip():
                paragrafos.append(paragrafo.text)

        return "\n".join(paragrafos), None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 📊 CSV
# ============================================================

def ler_csv(arquivo):
    """Lê um arquivo CSV."""

    try:
        conteudo = arquivo.read()

        if isinstance(conteudo, bytes):
            conteudo = conteudo.decode("utf-8", errors="ignore")

        linhas = []

        leitor = csv.reader(conteudo.splitlines())

        for linha in leitor:
            linhas.append(" | ".join(linha))

        return "\n".join(linhas), None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 📊 XLSX
# ============================================================

def ler_xlsx(arquivo):
    """Extrai dados de uma planilha XLSX."""

    if openpyxl is None:
        return None, "A biblioteca openpyxl não está instalada."

    try:
        dados = []

        planilha = openpyxl.load_workbook(
            BytesIO(arquivo.read()),
            read_only=True,
            data_only=True
        )

        for nome_planilha in planilha.sheetnames:

            folha = planilha[nome_planilha]

            dados.append(
                f"\n=== PLANILHA: {nome_planilha} ==="
            )

            for linha in folha.iter_rows(values_only=True):

                valores = []

                for valor in linha:

                    if valor is not None:
                        valores.append(str(valor))

                if valores:
                    dados.append(" | ".join(valores))

        return "\n".join(dados), None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 💻 CÓDIGO / TEXTO ESTRUTURADO
# ============================================================

def ler_codigo(arquivo):
    """Lê arquivos de código como texto."""

    try:
        conteudo = arquivo.read()

        if isinstance(conteudo, bytes):
            conteudo = conteudo.decode("utf-8", errors="ignore")

        return conteudo, None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# ⚙️ JSON
# ============================================================

def ler_json(arquivo):
    """Lê e organiza um arquivo JSON."""

    try:
        conteudo = arquivo.read()

        if isinstance(conteudo, bytes):
            conteudo = conteudo.decode("utf-8", errors="ignore")

        dados = json.loads(conteudo)

        return json.dumps(
            dados,
            indent=2,
            ensure_ascii=False
        ), None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 📦 ZIP INTELIGENTE
# ============================================================

def ler_zip(arquivo):
    """
    Analisa um arquivo ZIP.

    A função:
    - lista os arquivos internos;
    - identifica os formatos;
    - lê arquivos de texto/código compatíveis;
    - mantém o ZIP original intacto;
    - não extrai arquivos fisicamente para o servidor;
    - aplica limites de segurança.
    """

    try:

        dados = []

        total_lido = 0
        quantidade_processada = 0

        with zipfile.ZipFile(arquivo) as zip_arquivo:

            lista_arquivos = zip_arquivo.infolist()

            # ------------------------------------------------
            # 📋 RESUMO DO ZIP
            # ------------------------------------------------

            dados.append("=== 📦 ANÁLISE DO ARQUIVO ZIP ===")
            dados.append(
                f"Total de itens encontrados: {len(lista_arquivos)}"
            )

            dados.append("")
            dados.append("=== 📋 ESTRUTURA DO ZIP ===")

            for item in lista_arquivos:

                # Ignora diretórios.
                if item.is_dir():
                    continue

                tamanho = item.file_size

                dados.append(
                    f"- {item.filename} "
                    f"({tamanho} bytes)"
                )

            dados.append("")

            # ------------------------------------------------
            # 🛡️ LIMITE DE QUANTIDADE
            # ------------------------------------------------

            arquivos_processaveis = [
                item
                for item in lista_arquivos
                if not item.is_dir()
            ]

            if len(arquivos_processaveis) > MAX_ARQUIVOS_ZIP:

                dados.append(
                    "⚠️ Limite de segurança atingido."
                )

                dados.append(
                    f"O ZIP possui {len(arquivos_processaveis)} "
                    f"arquivos, mas somente os primeiros "
                    f"{MAX_ARQUIVOS_ZIP} serão analisados."
                )

            # ------------------------------------------------
            # 📖 LEITURA DOS ARQUIVOS INTERNOS
            # ------------------------------------------------

            dados.append("")
            dados.append(
                "=== 🧠 ARQUIVOS INTERNOS ANALISADOS ==="
            )

            for item in arquivos_processaveis:

                if quantidade_processada >= MAX_ARQUIVOS_ZIP:
                    break

                nome_interno = item.filename
                tamanho_interno = item.file_size

                extensao = Path(nome_interno).suffix.lower()

                # --------------------------------------------
                # 🛡️ ARQUIVO GRANDE DEMAIS
                # --------------------------------------------

                if tamanho_interno > MAX_ARQUIVO_INTERNO_ZIP:

                    dados.append("")
                    dados.append(
                        f"📄 ARQUIVO: {nome_interno}"
                    )

                    dados.append(
                        "⚠️ Arquivo ignorado: "
                        "tamanho interno acima do limite "
                        f"de {MAX_ARQUIVO_INTERNO_ZIP // (1024 * 1024)} MB."
                    )

                    continue

                # --------------------------------------------
                # 🛡️ LIMITE TOTAL
                # --------------------------------------------

                if (
                    total_lido + tamanho_interno
                    > MAX_TOTAL_ZIP
                ):

                    dados.append("")
                    dados.append(
                        f"📄 ARQUIVO: {nome_interno}"
                    )

                    dados.append(
                        "⚠️ Leitura interrompida: "
                        "limite total de dados do ZIP "
                        "atingido."
                    )

                    break

                # --------------------------------------------
                # 📦 ZIP INTERNO
                # --------------------------------------------

                if extensao == ".zip":

                    dados.append("")
                    dados.append(
                        f"📦 ARQUIVO: {nome_interno}"
                    )

                    dados.append(
                        "ℹ️ ZIP interno identificado. "
                        "Não será aberto nesta etapa."
                    )

                    quantidade_processada += 1
                    continue

                # --------------------------------------------
                # 📄 FORMATOS SUPORTADOS
                # --------------------------------------------

                extensoes_legiveis = (
                    ".txt",
                    ".py",
                    ".js",
                    ".ts",
                    ".html",
                    ".css",
                    ".java",
                    ".cpp",
                    ".c",
                    ".h",
                    ".sql",
                    ".sh",
                    ".md",
                    ".xml",
                    ".rtf",
                    ".json",
                    ".csv",
                )

                if extensao not in extensoes_legiveis:

                    dados.append("")
                    dados.append(
                        f"📄 ARQUIVO: {nome_interno}"
                    )

                    dados.append(
                        f"ℹ️ Formato {extensao or 'desconhecido'} "
                        "encontrado, mas a leitura automática "
                        "ainda não está disponível."
                    )

                    quantidade_processada += 1
                    continue

                # --------------------------------------------
                # 📖 LEITURA
                # --------------------------------------------

                try:

                    conteudo_bytes = zip_arquivo.read(item)

                    total_lido += len(conteudo_bytes)

                    arquivo_interno = BytesIO(
                        conteudo_bytes
                    )

                    # O leitor usa .name em alguns pontos.
                    arquivo_interno.name = nome_interno

                    # ----------------------------------------
                    # 🔀 ESCOLHE O LEITOR
                    # ----------------------------------------

                    if extensao == ".json":
                        conteudo, erro = ler_json(
                            arquivo_interno
                        )

                    elif extensao == ".csv":
                        conteudo, erro = ler_csv(
                            arquivo_interno
                        )

                    else:
                        conteudo, erro = ler_codigo(
                            arquivo_interno
                        )

                    # ----------------------------------------
                    # 📋 RESULTADO
                    # ----------------------------------------

                    dados.append("")
                    dados.append(
                        f"📄 ARQUIVO: {nome_interno}"
                    )

                    if erro:

                        dados.append(
                            f"❌ Erro ao ler arquivo: {erro}"
                        )

                    else:

                        if conteudo:

                            dados.append(
                                "```"
                            )

                            dados.append(
                                conteudo
                            )

                            dados.append(
                                "```"
                            )

                        else:

                            dados.append(
                                "ℹ️ Arquivo sem conteúdo."
                            )

                except Exception as erro_interno:

                    dados.append("")
                    dados.append(
                        f"📄 ARQUIVO: {nome_interno}"
                    )

                    dados.append(
                        "❌ Não foi possível ler o "
                        f"conteúdo: {erro_interno}"
                    )

                quantidade_processada += 1

        # ----------------------------------------------------
        # 📊 RESUMO FINAL
        # ----------------------------------------------------

        dados.append("")
        dados.append(
            "=== 📊 RESUMO DA ANÁLISE ==="
        )

        dados.append(
            f"Arquivos processados: "
            f"{quantidade_processada}"
        )

        dados.append(
            f"Dados lidos: "
            f"{round(total_lido / 1024, 2)} KB"
        )

        dados.append("")
        dados.append(
            "✅ O arquivo ZIP original não foi modificado."
        )

        return "\n".join(dados), None

    except zipfile.BadZipFile:
        return None, (
            "O arquivo enviado não é um ZIP válido "
            "ou está corrompido."
        )

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 📋 INFORMAÇÕES DO ARQUIVO
# ============================================================

def obter_info_arquivo(arquivo):
    """Retorna informações básicas do arquivo."""

    try:

        nome = arquivo.name
        tamanho = arquivo.size

        extensao = Path(nome).suffix.lower()

        return {
            "nome": nome,
            "tamanho_bytes": tamanho,
            "tamanho_kb": round(
                tamanho / 1024,
                2
            ),
            "extensao": extensao,
        }, None

    except Exception as erro:
        return None, str(erro)


# ============================================================
# 🧠 FUNÇÃO PRINCIPAL
# ============================================================

def ler_arquivo(arquivo):
    """
    Identifica o tipo do arquivo
    e extrai seu conteúdo.
    """

    if arquivo is None:
        return None, "Nenhum arquivo foi enviado."

    nome = arquivo.name.lower()

    # --------------------------------------------------------
    # 📄 DOCUMENTOS
    # --------------------------------------------------------

    if nome.endswith(".txt"):
        return ler_txt(arquivo)

    if nome.endswith(".pdf"):
        return ler_pdf(arquivo)

    if nome.endswith(".docx"):
        return ler_docx(arquivo)

    # --------------------------------------------------------
    # 📊 PLANILHAS
    # --------------------------------------------------------

    if nome.endswith(".csv"):
        return ler_csv(arquivo)

    if nome.endswith(".xlsx"):
        return ler_xlsx(arquivo)

    # --------------------------------------------------------
    # ⚙️ DADOS
    # --------------------------------------------------------

    if nome.endswith(".json"):
        return ler_json(arquivo)

    # --------------------------------------------------------
    # 💻 CÓDIGO
    # --------------------------------------------------------

    extensoes_codigo = (
        ".py",
        ".js",
        ".ts",
        ".html",
        ".css",
        ".java",
        ".cpp",
        ".c",
        ".h",
        ".sql",
        ".sh",
        ".md",
        ".xml",
        ".rtf",
    )

    if nome.endswith(extensoes_codigo):
        return ler_codigo(arquivo)

    # --------------------------------------------------------
    # 📦 ZIP
    # --------------------------------------------------------

    if nome.endswith(".zip"):
        return ler_zip(arquivo)

    # --------------------------------------------------------
    # 🖼️ / 🎵 / 🎬 MULTIMÍDIA
    # --------------------------------------------------------

    extensoes_multimidia = (
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".gif",
        ".bmp",
        ".tiff",
        ".tif",
        ".mp3",
        ".wav",
        ".ogg",
        ".m4a",
        ".flac",
        ".aac",
        ".mp4",
        ".mov",
        ".avi",
        ".mkv",
        ".webm",
        ".mpeg4",
    )

    if nome.endswith(extensoes_multimidia):

        return (
            "",
            None
        )

    # --------------------------------------------------------
    # ❌ FORMATO NÃO SUPORTADO
    # --------------------------------------------------------

    return None, (
        f"Formato não suportado: "
        f"{Path(nome).suffix}"
                        )
